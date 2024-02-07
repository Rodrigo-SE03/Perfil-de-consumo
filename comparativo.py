import pandas as pd
import equipamentos,consumo
from openpyxl import load_workbook
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.worksheet.merge import MergeCell
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, PatternFill, Alignment
import xlwings as xw
from time import sleep

def criar_complementar(itens,categoria,tarifas,values,xl_name):
    writer = pd.ExcelWriter(f'{xl_name.split(".")[0]}_temp.xlsx',engine="xlsxwriter")
    equipamentos.criar_equip(itens,writer)
    if categoria == 'Convencional':
        categoria = 'Branca'
        tarifas = [values['-tarifaFP_branca-'],values['-tarifaP_branca-'],values['-tarifaI_branca-']]
    elif categoria == 'Branca':
        categoria = 'Convencional'
        tarifas = [values['-tarifa_conv-']]
    elif categoria == 'Verde':
        categoria = 'Azul'
        tarifas = [values['-tarifaFP_azul-'],values['-tarifaP_azul-'],values['-tarifaDEMFP_azul-'],values['-tarifaDEMP_azul-']]
    else:
        categoria = 'Verde'
        tarifas = [values['-tarifaFP_verde-'],values['-tarifaP_verde-'],values['-tarifaDEM_verde-']]
    print(tarifas)
    consumo.criar_consumo(itens,writer,categoria,tarifas,values)
    writer.close()

def comparar(xl_name,categoria):

    borda = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    
    if categoria == 'Branca' or categoria == 'Convencional':
        with xw.App(visible=False) as app:
            wb_og = xw.Book(xl_name,read_only = True)
            ws_og = wb_og.sheets['Consumo geral']
            custo_mensal_og = ws_og['N2'].value if categoria == 'Convencional' else ws_og['O5'].value
            wb_og.save()
            wb_og.close()
            

        with xw.App(visible=False) as app: 
            wb_comp = xw.Book(f'{xl_name.split(".")[0]}_temp.xlsx')
            ws_comp = wb_comp.sheets['Consumo geral']
            custo_mensal_comp = ws_comp['O5'].value if categoria == 'Convencional' else ws_comp['N2'].value
            wb_comp.save()
            wb_comp.close()

        cat_comp = 'Convencional' if categoria == 'Branca' else 'Branca'

        
        diferenca = custo_mensal_comp - custo_mensal_og

        wb_og = load_workbook(xl_name)
        wb_og.create_sheet("Comparativo")
        wb_og.save(xl_name)


        ws_og = wb_og['Comparativo']
        ws_og.merge_cells('C3:D3')

        titulo = ws_og.cell(3,3)
        titulo.font = Font(color = '00FFFFFF',bold=True)
        titulo.fill = PatternFill('solid',start_color='4f81bd',end_color='4f81bd')
        titulo.value = 'Comparação de Custos'
        titulo.alignment = Alignment(horizontal='center',vertical='center')

        ws_og['C4'].value = cat_comp
        ws_og['D4'].value = custo_mensal_comp

        ws_og['C5'].value = categoria
        ws_og['D5'].value = custo_mensal_og

        ws_og['C6'].value = 'Diferença'
        ws_og['D6'].value = diferenca

        for i in range(3,7):
            for j in range(3,5):
                ws_og.cell(i,j).border = borda
                ws_og.cell(i,j).number_format = "R$ #,##0.00"
        ws_og.column_dimensions['C'].width = 20
        ws_og.column_dimensions['D'].width = 20

        wb_og.save(xl_name)


    else:
        with xw.App(visible=False) as app:
            wb_og = xw.Book(xl_name,read_only = True)
            ws_og = wb_og.sheets['Consumo geral']
            custo_mensal_og = ws_og['P6'].value if categoria == 'Azul' else ws_og['P5'].value
            ws_og = wb_og.sheets['Reativos']
            custo_mensal_og_r = ws_og['N4'].value if categoria == 'Verde' else ws_og['O5'].value
            wb_og.save()
            wb_og.close()
            

        with xw.App(visible=False) as app: 
            wb_comp = xw.Book(f'{xl_name.split(".")[0]}_temp.xlsx')
            ws_comp = wb_comp.sheets['Consumo geral']
            custo_mensal_comp = ws_comp['P5'].value if categoria == 'Azul' else ws_comp['P6'].value
            ws_comp = wb_comp.sheets['Reativos']
            custo_mensal_comp_r = ws_comp['O5'].value if categoria == 'Verde' else ws_comp['N4'].value
            wb_comp.save()
            wb_comp.close()
        
        cat_comp = 'Azul' if categoria == 'Verde' else 'Verde'

        diff_a = custo_mensal_comp - custo_mensal_og
        diff_r = custo_mensal_comp_r - custo_mensal_og_r
        
        wb_og = load_workbook(xl_name)
        wb_og.create_sheet("Comparativo")
        wb_og.save(xl_name)

        ws_og = wb_og['Comparativo']

        ws_og.merge_cells('C3:D3')
        titulo_a = ws_og.cell(3,3)
        titulo_a.font = Font(color = '00FFFFFF',bold=True)
        titulo_a.fill = PatternFill('solid',start_color='4f81bd',end_color='4f81bd')
        titulo_a.value = 'Comparação de Custos com Energia Ativa'
        titulo_a.alignment = Alignment(horizontal='center',vertical='center')

        ws_og['C4'].value = cat_comp
        ws_og['D4'].value = custo_mensal_comp
        ws_og['C5'].value = categoria
        ws_og['D5'].value = custo_mensal_og
        ws_og['C6'].value = 'Diferença'
        ws_og['D6'].value = diff_a

        for i in range(3,7):
            for j in range(3,5):
                ws_og.cell(i,j).border = borda
                ws_og.cell(i,j).number_format = "R$ #,##0.00"

        ws_og.merge_cells('C9:D9')
        titulo_r = ws_og.cell(9,3)
        titulo_r.font = Font(color = '00FFFFFF',bold=True)
        titulo_r.fill = PatternFill('solid',start_color='4f81bd',end_color='4f81bd')
        titulo_r.value = 'Comparação de Custos com Reativos'
        titulo_r.alignment = Alignment(horizontal='center',vertical='center')

        ws_og['C10'].value = cat_comp
        ws_og['D10'].value = custo_mensal_comp_r
        ws_og['C11'].value = categoria
        ws_og['D11'].value = custo_mensal_og_r
        ws_og['C12'].value = 'Diferença'
        ws_og['D12'].value = diff_r

        for i in range(9,13):
            for j in range(3,5):
                ws_og.cell(i,j).border = borda
                ws_og.cell(i,j).number_format = "R$ #,##0.00"

        ws_og.column_dimensions['C'].width = 20
        ws_og.column_dimensions['D'].width = 20

        wb_og.save(xl_name)